import openpyxl
import shutil
import os
import datetime
'''
This program to be run after running this query to automatically generate the SQL needed to update the info


--aqt export,file=Email_updates.xlsx,type=excel,sheet=EMAILS,filemode=replace,prompt=no,header=yes,nulls=blank,sdelim=dquote
SELECT DISTINCT A.EMPLID, A.E_ADDR_TYPE, A.EMAIL_ADDR AS CS_BUSN, B.EMAIL_ADDR AS HR_BUSN
FROM PS_EMAIL_ADDRESSES A, PS_EMAIL_ADDRESSES@HR_PUB B, PS_EMPLOYEES@HR_PUB C
WHERE A.EMPLID = B.EMPLID  
AND A.EMPLID = C.EMPLID
AND C.ANNUAL_RT > 0
AND A.E_ADDR_TYPE = B.E_ADDR_TYPE
AND A.E_ADDR_TYPE = 'BUSN'
AND A.EMAIL_ADDR <> B.EMAIL_ADDR
AND UPPER(A.EMAIL_ADDR) LIKE '%UFL.EDU%'
AND UPPER(B.EMAIL_ADDR) NOT LIKE '%UFL.EDU%';


'''

def build_BUSN_queries(BUSN):
       
    with open (BUSN, "w") as file:
        import openpyxl
        wb = openpyxl.load_workbook(filename='Email_updates.xlsx', read_only=False)
        sheet = wb.get_sheet_by_name('EMAILS')
        for row in range(2, sheet.max_row + 1):
            EMPL_ID = sheet['A' + str(row)].value
            BUSN_EMAIL = sheet['C' + str(row)].value
            PERS_EMAIL = sheet['D' + str(row)].value
            file.write("--1 row" + "\n" + "UPDATE PS_EMAIL_ADDRESSES SET EMAIL_ADDR = " + "'" + BUSN_EMAIL + "'" + "\n" + "WHERE E_ADDR_TYPE = " + "'" + "BUSN" + "'" + "\n" + "AND EMPLID = " + "'" + EMPL_ID + "';" + "\n" + "\n")
    

def build_PERS_queries(PERS):

    with open (PERS, "w") as file:
        import openpyxl
        wb = openpyxl.load_workbook(filename='Email_updates.xlsx', read_only=False)
        sheet = wb.get_sheet_by_name('EMAILS')
        for row in range(2, sheet.max_row + 1):
            EMPL_ID = sheet['A' + str(row)].value
            BUSN_EMAIL = sheet['C' + str(row)].value
            PERS_EMAIL = sheet['D' + str(row)].value
            file.write("--1 row" + "\n" + "INSERT INTO PS_EMAIL_ADDRESSES VALUES ('" + EMPL_ID + "','PERS','" + PERS_EMAIL + "','N');" + "\n" + "\n")

'''
After running this query, the following function builds the inserts

--aqt export,file=Email_updates_ufl_email.xlsx,type=excel,sheet=EMAILS,filemode=replace,prompt=no,header=yes,nulls=blank,sdelim=dquote
SELECT   Distinct A.EMPLID,
         B.E_ADDR_TYPE,
         B.EMAIL_ADDR,
         'Y'
FROM     PS_EMPLOYEES@HR_PUB A,
         PS_EMAIL_ADDRESSES B
WHERE    A.EMPLID = B.EMPLID
AND      B.E_ADDR_TYPE = 'BUSN'
AND      A.ANNUAL_RT > 0
AND      UPPER(B.EMAIL_ADDR) LIKE '%UFL.EDU%'
AND      A.EMPLID NOT IN (SELECT C.EMPLID FROM PS_EMAIL_ADDRESSES@HR_PUB C WHERE A.EMPLID = C.EMPLID
     AND C.E_ADDR_TYPE = 'BUSN') 

'''

def build_atUFL_list_query(UFL_SQL):

    with open (UFL_SQL, "w") as file:
        file.write("--Run this is HRPRD to see if there are BUSN emails for these people \n " + "\n" + "SELECT * FROM PS_EMAIL_ADDRESSES WHERE E_ADDR_TYPE = 'BUSN' AND EMPLID IN (")
        wb = openpyxl.load_workbook(filename='Email_updates_ufl_email.xlsx', read_only=False)
        sheet = wb.get_sheet_by_name('EMAILS')
        for row in range(2, sheet.max_row +1 ):
            EMPL_ID = sheet['A' + str(row)].value
            file.write("'" + EMPL_ID + "', ")
        file.write(")")

'''
^^^ RUN THAT QUERY TO SEE IF HR HAS a BUSN EMAIL FOR THIS PERSON, IF THEY DO, EDIT AND REMOVE THOSE EMPLIDs
'''

def build_atUFL_insert(UFL):
    
    with open (UFL, "w") as file:
        wb = openpyxl.load_workbook(filename='Email_updates_ufl_email.xlsx', read_only=False)
        sheet = wb.get_sheet_by_name('EMAILS')
        for row in range(2, sheet.max_row +1 ):
            EMPL_ID = sheet['A' + str(row)].value
            BUSN = sheet['B' + str(row)].value
            EMAIL = sheet['C' + str(row)].value
            FLAG = sheet['D' + str(row)].value
            file.write("INSERT into PS_EMAIL_ADDRESSES values " + " ('" + EMPL_ID + "', " + "'" + BUSN + "', " + "'" + EMAIL + "', '" + FLAG + "');\n")
        
    

def rename_files(BUSN, PERS, UFL_SQL, UFL):

    SR_Number = input("Please input the service request for this email sync update: ")
    Current_Date = datetime.datetime.today().strftime ('%d-%b-%Y')
    os.rename(r'BUSN_EMAIL_SYNC.txt',r'BUSN_EMAIL_SYNC_' + str(Current_Date) + "_" + "SR" + SR_Number + '.txt')
    print(" . ")
    os.rename(r'PERS_EMAIL_SYNC.txt',r'PERS_EMAIL_SYNC_' + str(Current_Date) + "_" + "SR" + SR_Number + '.txt')
    print(" . ")
    os.rename(r'UFL_EMAIL_SYNC_SQL.txt',r'UFL_EMAIL_SYNC_SQL_' + str(Current_Date) + "_" + "SR" + SR_Number + '.txt')
    print(" . ")
    os.rename(r'UFL_EMAIL_INSERTS.txt',r'UFL_EMAIL_INSERTS_' + str(Current_Date) + "_" + "SR" + SR_Number + '.txt')
    print("Done")
    


def main():
    os.chdir('H:\SavedQueries')
    BUSN = "BUSN_EMAIL_SYNC.txt"
    PERS = "PERS_EMAIL_SYNC.txt"
    UFL_SQL = "UFL_EMAIL_SYNC_SQL.txt"
    UFL = "UFL_EMAIL_INSERTS.txt"
    build_BUSN_queries(BUSN)
    build_PERS_queries(PERS)
    build_atUFL_list_query(UFL_SQL)
    build_atUFL_insert(UFL)
    rename_files(BUSN, PERS, UFL_SQL, UFL)
    


if __name__ == "__main__":
    main()
